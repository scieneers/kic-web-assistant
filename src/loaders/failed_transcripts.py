from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from src.loaders.models.module import Module
from src.loaders.models.moodlecourse import MoodleCourse


class FailedModule(BaseModel):
    modul: Module
    err_message: str


class FailedCourse(BaseModel):
    course: MoodleCourse
    modules: list[FailedModule]


class FailedTranscripts(BaseModel):
    courses: list[FailedCourse]


def save_failed_transcripts_to_excel(transcripts: FailedTranscripts, file_name: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fehlgeschlagene Transcripte"

    # Define headers
    headers = ["Kurs ID", "Kursname", "Module ID", "Modulename", "Modul-Link", "Fehlermeldung"]
    ws.append(headers)

    # Populate rows with data
    for course in transcripts.courses:
        for module in course.modules:
            row = [
                course.course.id,
                course.course.fullname,
                module.modul.id,
                module.modul.name,
                str(module.modul.url),
                module.err_message,
            ]
            ws.append(row)

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 20

    # Save the workbook
    wb.save(file_name)
    print(f"Data has been saved to {file_name}")
